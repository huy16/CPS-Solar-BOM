import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def create_template():
    output_file = "BOQ_BOM_Template.xlsx"
    
    wb = Workbook()
    
    # ---------------------------------------------------------
    # 1. Create BOQ Sheet
    # ---------------------------------------------------------
    ws_boq = wb.active
    ws_boq.title = "BOQ"
    
    # Define headers based on reference_vba_supplement.bas
    # Columns 1-3 are generic
    # Columns 4-23 are specific material keys
    
    headers_boq = [
        "No.", "Category / Item Name", "Unit", 
        "ER-R-ECO/5200",    # Col 4
        "ER-SP-ECO",        # Col 5
        "I-05-6.3/85/CS",   # Col 6
        "ER-EC-ST35",       # Col 7
        "ER-IC-ST35",       # Col 8
        "EZ-CC-PV/4",       # Col 9
        "EZ-GC-ST",         # Col 10
        "EZ-GL-ST",         # Col 11
        "CS7N-665MS",       # Col 12
        "T4-T6",            # Col 13
        "SUN2000",          # Col 14
        "Sdongle5A",        # Col 15
        "DTSU",             # Col 16
        "CXV-3x16-10sqrt",  # Col 17
        "C10sqrt",          # Col 18
        "Leader",           # Col 19
        "LS",               # Col 20
        "1/2-3/4",          # Col 21
        "JUNC_BOX",         # Col 22
        "SUNTREE"           # Col 23
    ]
    
    ws_boq.append(headers_boq)
    
    # Style BOQ Header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for cell in ws_boq[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        
    # Adjust column widths for BOQ
    ws_boq.column_dimensions['A'].width = 5
    ws_boq.column_dimensions['B'].width = 40
    ws_boq.column_dimensions['C'].width = 10
    for col_idx in range(4, len(headers_boq) + 1):
        col_letter = ws_boq.cell(row=1, column=col_idx).column_letter
        ws_boq.column_dimensions[col_letter].width = 15

    # ---------------------------------------------------------
    # 2. Create BOM Sheet
    # ---------------------------------------------------------
    ws_bom = wb.create_sheet("BOM")
    
    headers_bom = [
        "No.", "Category", "Item Name", "Specification", "Unit", "Quantity", "Note"
    ]
    
    ws_bom.append(headers_bom)
    
    # Style BOM Header
    for cell in ws_bom[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        
    # Adjust column widths for BOM
    ws_bom.column_dimensions['A'].width = 5
    ws_bom.column_dimensions['B'].width = 20
    ws_bom.column_dimensions['C'].width = 30
    ws_bom.column_dimensions['D'].width = 40
    ws_bom.column_dimensions['E'].width = 10
    ws_bom.column_dimensions['F'].width = 15
    ws_bom.column_dimensions['G'].width = 30

    # Save
    wb.save(output_file)
    print(f"Successfully created '{output_file}'")

if __name__ == "__main__":
    create_template()
